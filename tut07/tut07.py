from datetime import datetime
start_time = datetime.now()


def octant_analysis(mod=5000):
	try:
		import pandas as pd
		import numpy as np
		import math
		import glob
		import xlsxwriter
		import os
		import shutil
		from openpyxl import Workbook
		import openpyxl
		from openpyxl.styles import PatternFill
		from openpyxl.styles import Border,Side
		from datetime import datetime
		pd.io.formats.excel.ExcelFormatter.header_style = None
	except:
		print("import packages")
	check = mod
	# df = pd.read_excel("1.0.xlsx")
	x = os.path.realpath(os.path.dirname(__file__))
	print(x)
	path = x+'\input'
	path_out = x+'\output'
	os.chdir(path)
	my_files = glob.glob('*.xlsx')
	sat=0
	# roll_sheet.to_excel(f'{ip["Roll No"][sno]}.xlsx', index=False)
	for i in my_files :
	    try:
	        os.chdir(path)
	        df=pd.read_excel(i)

	        df.loc[0,"U_Avg"] = df["U"].mean()       #Creating average for coloumns U,V,W

	        df.loc[0,"V_Avg"] = df["V"].mean()

	        df.loc[0,"W_Avg"] = df["W"].mean()

	        df["U1"] = df["U"]-df["U"].mean()   # Creating new columns for U1,V2,W3     =
	        df["V2"] = df["V"]-df["V"].mean()
	        df["W3"] = df["W"]- df["W"].mean()
	        df.U_Avg=df.U_Avg.round(3)
	        df.V_Avg=df.V_Avg.round(3)
	        df.W_Avg=df.W_Avg.round(3)
	        df.U = df.U.round(3)
	        df.V = df.V.round(3)
	        df.W = df.W.round(3)
	        df.loc[((df.U1 >= 0) & (df.V2 >= 0) & (df.W3 >=0)), "Octant"] = "+1"
	        df.loc[((df.U1 > 0) &(df.V2 > 0) & (df.W3 <0)), "Octant" ] = "-1"
	        df.loc[((df.U1 < 0) &(df.V2 > 0) & (df.W3 >0)), "Octant" ] = "+2"
	        df.loc[((df.U1 < 0) &(df.V2 > 0) & (df.W3 <0)), "Octant" ] = "-2"    #creating octant column,   assigning integers for each octant
	        df.loc[((df.U1 < 0) &(df.V2 < 0) & (df.W3 >0)), "Octant" ] = "+3"
	        df.loc[((df.U1 < 0) &(df.V2 < 0) & (df.W3 <0)), "Octant" ] = "-3"
	        df.loc[((df.U1 > 0) &(df.V2 < 0) & (df.W3 >0)), "Octant" ] = "+4"
	        df.loc[((df.U1 > 0) &(df.V2 < 0) & (df.W3 <0)), "Octant" ] = "-4"
	        df.loc[0,"     "]=" "

	        # x =df['Octant'].value_counts() # total count of number of values for each octane
	        # lis= []
	        # for num in range(len(x)):
	        #   lis.append(num)
	        # lis = np.array(x)
	        # print(lis)
	        # exit()
	        df.loc[1,""]="Mod 5000"
	        df.loc[0,"octant id"]="overall count"  # creating octant id column and assigning octant count under that.
	        oct_count = df['Octant'].value_counts()
	        arr = ["+1", "-1", "+2", "-2", "+3", "-3",
	                "+4", "-4"]  # cretaed for reference
	        oct_cnt = {}  # for storing octant count as key and and coreesponding octant value as value in dict
	        for i in range(8):
	            s = arr[i]
	            # appending the overall count of octant and octant value in dict i.e for Ex:(2610,"+1")
	            oct_cnt.update({s:oct_count[s]})
	            # And assigning a count values to respectively Coloumns
	            df.loc[0, s] = oct_count[s]
	        # for i in ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]:
	        #     if i in x[i]:
	        #         df.loc[0, "{}".format(i)] = x["{}".format(i)]
	        #     else:
	        #         df.loc[0, "{}".format(i)] = 0

	        # df.loc[0,"+1"]=lis[0]      # assigning octant count under all octants(+1,-1,+2,-2,+3,-3,+4,-4)
	        # df.loc[0,"-1"]=lis[1]
	        # df.loc[0,"+2"]=lis[2]
	        # df.loc[0,"-2"]=lis[3]
	        # df.loc[0,"+3"]=lis[4]
	        # df.loc[0,"-3"]=lis[5]
	        # df.loc[0,"+4"]=lis[6]
	        # df.loc[0,"-4"]=lis[7]
	        mod=5000
	        y=str(mod)
	        d=math.ceil(19997/mod) # greatest integer function for identifing
	        l=0000
	        m=mod-1
	        a=str(l)
	        b=str(m)
	        for j in range(d) :
	           if int(b)>=19996:
	              df.loc[j+1,"octant id"]= a+"-"+"19996"
	              l=m+1
	              m=m+mod
	              a=str(l)
	              b=str(m)
	           else:
	              df.loc[j+1,"octant id"]= a+"-"+b
	              l=m+1
	              m=m+mod
	              a=str(l)
	              b=str(m)
	        c=0
	        t=0
	        p=1
	        for j in range(d) : # no of coloumns in the output for each octant
	            for i in range(mod) : #running at each 5000 iterations (0-5000,5001-10000,......25000-30000)
	                if df["Octant"][t]=="+4" : # counting number of +4 octant in range of 0-30000
	                   c =c+1
	                t=t+1
	                if t== 19997 : #we have break loop at t=29745 because after
	                  break
	            df.loc[p,"+4"]=c #assigning count of +4 in each coloumn by iterating p
	            p=p+1
	            c=0
	            j=7
	        c=0
	        t=0
	        p=1
	        for j in range(d) :# no of coloumns in the output for each octant
	            for i in range(mod) :
	                if df["Octant"][t]=="-4" :# counting number of -4 octant in range of 0-30000
	                   c =c+1
	                t=t+1
	                if t== 19997 :
	                  break
	            df.loc[p,"-4"]=c #assigning count of -4 in each coloumn by iterating p
	            p=p+1
	            c=0
	            j=7
	        c=0
	        t=0
	        p=1
	        for j in range(d) :# no of coloumns in the output for each octant
	            for i in range(mod) :
	                if df["Octant"][t]=="+3" :# counting number of +3 octant in range of 0-30000
	                   c =c+1
	                t=t+1
	                if t== 19997 :
	                  break
	            df.loc[p,"+3"]=c #assigning count of +3 in each coloumn by iterating p
	            p=p+1
	            c=0
	            j=7
	        c=0
	        t=0
	        p=1
	        for j in range(d) :# no of coloumns in the output for each octant
	            for i in range(mod) :
	                if df["Octant"][t]=="-3" :# counting number of -3 octant in range of 0-30000
	                   c =c+1
	                t=t+1
	                if t== 19997 :
	                  break
	            df.loc[p,"-3"]=c #assigning count of -3 in each coloumn by iterating p
	            p=p+1
	            c=0
	            j=7
	        c=0
	        t=0
	        p=1
	        for j in range(d) :# no of coloumns in the output for each octant
	            for i in range(mod) :
	                if df["Octant"][t]=="+2" :# counting number of +2 octant in range of 0-30000
	                   c =c+1
	                t=t+1
	                if t== 19997 :
	                  break
	            df.loc[p,"+2"]=c #assigning count of +2 in each coloumn by iterating p
	            p=p+1
	            c=0
	            j=7
	        c=0
	        t=0
	        p=1
	        for j in range(d) :# no of coloumns in the output for each octant
	            for i in range(mod) :
	                if df["Octant"][t]=="-2" :# counting number of -2 octant in range of 0-30000
	                   c =c+1
	                t=t+1
	                if t== 19997 :
	                  break
	            df.loc[p,"-2"]=c #assigning count of -2 in each coloumn by iterating p
	            p=p+1
	            c=0
	            j=7
	        c=0
	        t=0
	        p=1
	        for j in range(d) :# no of coloumns in the output for each octant
	            for i in range(mod) :
	                if df["Octant"][t]=="+1" :# counting number of +1 octant in range of 0-30000
	                   c =c+1
	                t=t+1
	                if t== 19997 :
	                  break
	            df.loc[p,"+1"]=c#assigning count of +1 in each coloumn by iterating p
	            p=p+1
	            c=0
	            j=7
	        c=0
	        t=0
	        p=1
	        for j in range(d) :# no of coloumns in the output for each octant
	            for i in range(mod) :
	                if df["Octant"][t]=="-1" :# counting number of -1 octant in range of 0-30000
	                   c =c+1
	                t=t+1
	                if t== 19997 :
	                  break
	            df.loc[p,"-1"]=c #assigning count of -1 in each coloumn by iterating p
	            p=p+1
	            c=0
	            j=7
	        a=[]
	        b=[]
	        c=[]
	        for i in range(8) :
	            c.append(0)
	        a.append(df["+1"][0])
	        a.append(df["-1"][0])
	        a.append(df["+2"][0])
	        a.append(df["-2"][0])
	        a.append(df["+3"][0])
	        a.append(df["-3"][0])
	        a.append(df["+4"][0])
	        a.append(df["-4"][0])
	        b.append(df["+1"][0])
	        b.append(df["-1"][0])
	        b.append(df["+2"][0])
	        b.append(df["-2"][0])
	        b.append(df["+3"][0])
	        b.append(df["-3"][0])
	        b.append(df["+4"][0])
	        b.append(df["-4"][0])
	        b.sort(reverse = True)
	        for i in range(8) :
	            for j in range(8) :
	                if a[j]==b[i] :
	                    c[j]=i+1
	                    j=9
	        df.loc[0,"Rank Octant 1"]=c[0]
	        df.loc[0,"Rank Octant -1"]=c[1]
	        df.loc[0,"Rank Octant 2"]=c[2]
	        df.loc[0,"Rank Octant -2"]=c[3]
	        df.loc[0,"Rank Octant 3"]=c[4]
	        df.loc[0,"Rank Octant -3"]=c[5]
	        df.loc[0,"Rank Octant 4"]=c[6]
	        df.loc[0,"Rank Octant -4"]=c[7]
	        for v in range(d) :
	         e=[]
	         f=[]
	         g=[]
	         for q in range(8) :
	          g.append(0)
	         e.append(df["+1"][v+1])
	         e.append(df["-1"][v+1])
	         e.append(df["+2"][v+1])
	         e.append(df["-2"][v+1])
	         e.append(df["+3"][v+1])
	         e.append(df["-3"][v+1])
	         e.append(df["+4"][v+1])
	         e.append(df["-4"][v+1])
	         f.append(df["+1"][v+1])
	         f.append(df["-1"][v+1])
	         f.append(df["+2"][v+1])
	         f.append(df["-2"][v+1])
	         f.append(df["+3"][v+1])
	         f.append(df["-3"][v+1])
	         f.append(df["+4"][v+1])
	         f.append(df["-4"][v+1])
	         f.sort(reverse = True)
	         for i in range(8) :
	              for j in range(8) :
	                  if e[j]==f[i] :
	                      g[j]=i+1
	                      j=9
	         df.loc[v+1,"Rank Octant 1"]=g[0]
	         df.loc[v+1,"Rank Octant -1"]=g[1]
	         df.loc[v+1,"Rank Octant 2"]=g[2]
	         df.loc[v+1,"Rank Octant -2"]=g[3]
	         df.loc[v+1,"Rank Octant 3"]=g[4]
	         df.loc[v+1,"Rank Octant -3"]=g[5]
	         df.loc[v+1,"Rank Octant 4"]=g[6]
	         df.loc[v+1,"Rank Octant -4"]=g[7]
	         w=[]
	         c=0
	         w.append(df["Rank Octant 1"][0])
	         w.append(df["Rank Octant -1"][0])
	         w.append(df["Rank Octant 2"][0])
	         w.append(df["Rank Octant -2"][0])
	         w.append(df["Rank Octant 3"][0])
	         w.append(df["Rank Octant -3"][0])
	         w.append(df["Rank Octant 4"][0])
	         w.append(df["Rank Octant -4"][0])

	         for j in range(8) :
	             if w[j]==1 :
	                 c=j
	                 if(c==0) :
	                    c=1
	                    df.loc[0,"Rank1 Octant ID"]=c
	                    j=8
	                 if(c==1) :
	                    c=-1
	                    df.loc[0,"Rank1 Octant ID"]=c
	                    j=8
	                 if(c==2) :
	                    c=2
	                    df.loc[0,"Rank1 Octant ID"]=c
	                    j=8
	                 if(c==3) :
	                    c=-2
	                    df.loc[0,"Rank1 Octant ID"]=c
	                    j=8
	                 if(c==4) :
	                     c=3
	                     df.loc[0,"Rank1 Octant ID"]=c
	                     j=8
	                 if(c==5) :
	                     c=-3
	                     df.loc[0,"Rank1 Octant ID"]=c
	                     j=8
	                 if(c==6) :
	                     c=4
	                     df.loc[0,"Rank1 Octant ID"]=c
	                     j=8
	                 if(c==7) :
	                    c=-4
	                    df.loc[0,"Rank1 Octant ID"]=c
	                    j=8
	         for i in range(d) :
	            y=[]
	            c=0
	            y.append(df["Rank Octant 1"][i+1])
	            y.append(df["Rank Octant -1"][i+1])
	            y.append(df["Rank Octant 2"][i+1])
	            y.append(df["Rank Octant -2"][i+1])
	            y.append(df["Rank Octant 3"][i+1])
	            y.append(df["Rank Octant -3"][i+1])
	            y.append(df["Rank Octant 4"][i+1])
	            y.append(df["Rank Octant -4"][i+1])
	            for j in range(8) :
	                if y[j]==1 :
	                    c=j
	                    if(c==0) :
	                       c=1
	                       df.loc[i+1,"Rank1 Octant ID"]=c
	                       j=8
	                    if(c==1) :
	                       c=-1
	                       df.loc[i+1,"Rank1 Octant ID"]=c
	                       j=8
	                    if(c==2) :
	                       c=2
	                       df.loc[i+1,"Rank1 Octant ID"]=c
	                       j=8
	                    if(c==3) :
	                       c=-2
	                       df.loc[i+1,"Rank1 Octant ID"]=c
	                       j=8
	                    if(c==4) :
	                        c=3
	                        df.loc[i+1,"Rank1 Octant ID"]=c
	                        j=8
	                    if(c==5) :
	                        c=-3
	                        df.loc[i+1,"Rank1 Octant ID"]=c
	                        j=8
	                    if(c==6) :
	                        c=4
	                        df.loc[i+1,"Rank1 Octant ID"]=c
	                        j=8
	                    if(c==7) :
	                       c=-4
	                       df.loc[i+1,"Rank1 Octant ID"]=c
	                       j=8

	        df.loc[d+5,"-1"]="OCTANT ID"
	        k=0
	        o=0
	        n=d+6
	        for i in range(8) :
	            if(i%2==0) :
	                k=k+1
	                df.loc[n,"-1"]=k
	                n=n+1
	            else :
	                o=o-1
	                df.loc[n,"-1"]=o
	                n=n+1

	        octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
	                                      "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}
	        df.loc[d+5,"+2"]="Octant Name "
	        df.loc[d+5,"-2"]="Count of Rank 1 Mod Values"
	        k=0
	        o=0
	        n=d+6
	        for i in range(8) :
	            if(i%2==0) :
	                k=k+1
	                df.loc[n,"+2"]=octant_name_id_mapping[str(k)]
	                n=n+1
	            else :
	                o=o-1
	                df.loc[n,"+2"]=octant_name_id_mapping[str(o)]
	                n=n+1
	        df.loc[0,"Rank1 Octant Name"]=octant_name_id_mapping[str(int(df.loc[0,"Rank1 Octant ID"]))]
	        for i in range(d) :
	            df.loc[i+1,"Rank1 Octant Name"]=octant_name_id_mapping[str(int(df.loc[i+1,"Rank1 Octant ID"]))]
	        p=df['Rank1 Octant ID'].value_counts()
	        h=[]
	        for i in range(d) :
	            h.append(int(df["Rank1 Octant ID"][i+1]))
	        k=0
	        for i in range(d) :
	            if(h[i]==1)  :
	                k=k+1
	        df.loc[d+6,"-2"]=k
	        k=0
	        for i in range(d) :
	            if(h[i]==-1)  :
	                k=k+1
	        df.loc[d+7,"-2"]=k
	        k=0
	        for i in range(d) :
	            if(h[i]==2)  :
	                k=k+1
	        df.loc[d+8,"-2"]=k
	        k=0
	        for i in range(d) :
	            if(h[i]==-2)  :
	                k=k+1
	        df.loc[d+9,"-2"]=k
	        k=0
	        for i in range(d) :
	            if(h[i]==3)  :
	                k=k+1
	        df.loc[d+10,"-2"]=k
	        k=0
	        for i in range(d) :
	            if(h[i]==-3)  :
	                k=k+1
	        df.loc[d+11,"-2"]=k
	        k=0
	        for i in range(d) :
	            if(h[i]==4)  :
	                k=k+1
	        df.loc[d+12,"-2"]=k
	        k=0
	        for i in range(d) :
	            if(h[i]==-4)  :
	                k=k+1
	        df.loc[d+13,"-2"]=k
	        df.loc[0,"             "]=""
	        df.loc[0," "]="From"
	        df.loc[0,"Octant #"]=""
	        df.loc[0,"+1 "]=""
	        df.loc[0,"-1 "]=""
	        df.loc[0,"+2 "]=""   #assigning (+4 to -4) both rows and columns
	        df.loc[0,"-2 "]=""
	        df.loc[0,"+3 "]=""
	        df.loc[0,"-3 "]=""
	        df.loc[0,"+4 "]=""
	        df.loc[0,"-4 "]=""
	        bhanu=13
	        for i in range(d):
	            df.loc[bhanu-2,"Octant #"]="Mod Transition Count"
	            df.loc[bhanu-1,"+1 "]="To"
	            df.loc[bhanu,"Octant #"]="Octant #"
	            df.loc[bhanu,"+1 "]="+1"
	            df.loc[bhanu,"-1 "]="-1"
	            df.loc[bhanu,"+2 "]="+2"
	            df.loc[bhanu,"-2 "]="-2"
	            df.loc[bhanu,"+3 "]="+3"
	            df.loc[bhanu,"-3 "]="-3"
	            df.loc[bhanu,"+4 "]="+4"
	            df.loc[bhanu,"-4 "]="-4"
	            df.loc[bhanu+1," "]="From"
	            bhanu=bhanu+13
	        bhanu=14
	        for i in range(d):
	            df.loc[bhanu,"Octant #"]="+1"
	            df.loc[bhanu+1,"Octant #"]="-1"
	            df.loc[bhanu+2,"Octant #"]="+2"
	            df.loc[bhanu+3,"Octant #"]="-2"
	            df.loc[bhanu+4,"Octant #"]="+3"
	            df.loc[bhanu+5,"Octant #"]="-3"
	            df.loc[bhanu+6,"Octant #"]="+4"
	            df.loc[bhanu+7,"Octant #"]="-4"
	            bhanu=bhanu+13
	        df.loc[0,"Octant #"]="+1"
	        df.loc[1,"Octant #"]="-1"
	        df.loc[2,"Octant #"]="+2"
	        df.loc[3,"Octant #"]="-2"
	        df.loc[4,"Octant #"]="+3"
	        df.loc[5,"Octant #"]="-3"
	        df.loc[6,"Octant #"]="+4"
	        df.loc[7,"Octant #"]="-4"
	        for i in range(8) :
	            df.loc[i,"+1 "]=0
	            df.loc[i,"-1 "]=0
	            df.loc[i,"+2 "]=0
	            df.loc[i,"-2 "]=0
	            df.loc[i,"+3 "]=0
	            df.loc[i,"-3 "]=0
	            df.loc[i,"+4 "]=0
	            df.loc[i,"-4 "]=0
	        l=0000
	        m=mod-1
	        a=str(l)
	        b=str(m)
	        by=12
	        for j in range(d) :
	          if int(b)>=19996:
	           df.loc[by,"Octant #"]= a+"-"+"19996"  #assigning range  from total count(0-4999 ......25000-29999 ) if mode is 5000
	           l=m+1
	           m=m+mod
	           a=str(l)
	           b=str(m)
	           by=by+13
	          else:
	           df.loc[by,"Octant #"]= a+"-"+b  #assigning range  from total count(0-4999 ......25000-29999 ) if mode is 5000
	           l=m+1
	           m=m+mod
	           a=str(l)
	           b=str(m)
	           by=by+13

	        rows, cols = (9, 9) # assigning rows and column variables
	        arr = [[0 for i in range(cols)] for j in range(rows)] # creating array with all zeros with 9 rows 9 columns
	        for bh in range(19996) :
	            arr[int(df["Octant"][bh])+4][int(df["Octant"][bh+1])+4]=arr[int(df["Octant"][bh])+4][int(df["Octant"][bh+1])+4]+1 #count of all ranges is stored in array ,we add +4 to avoid negative index in array
	        z=5
	        for i in range(8) :
	            if(i%2==0) :
	              df.loc[i,"+1 "]=arr[z][5]               # count of +1 - +1 is stored in arr[4+1][4+1]   [adding 4 to each index +1,-1,+2,-2,+3,-3,+4,-4]
	            if(i%2!=0) :                             # count of +1 - -1 is stored in arr[4+1][4-1]
	               z=8-z                                 # count of +1 - +2 is stored in arr[4+1][2+4]
	               df.loc[i,"+1 "]=arr[z][5]              # count of +1 - -2 is stored in arr[1+4][-2+4]
	               z=8-z                                 # count of +1 - +3 is stored in arr[4+1][3+4]
	               z=z+1                                 # count of +1 - -3 is stored in arr[4+1][-3+4]
	        z=5                                          # count of +1 - +4 is stored in arr[4+1][4+4]
	        for i in range(8) :                          # count of +1 - -4 is stored in arr[4+1][-4+4]
	            if(i%2==0) :
	              df.loc[i,"-1 "]=arr[z][3]
	            if(i%2!=0) :                             # count of +2 - +1 is stored in arr[4+2][1+4]
	               z=8-z                                 # count of +2 - -1 is stored in arr[4+2][-1+4]
	               df.loc[i,"-1 "]=arr[z][3]              # count of +2 - +2 is stored in arr[4+2][2+4]
	               z=8-z                                 # count of +2 - -2 is stored in arr[4+2][-2+4]
	               z=z+1                                 # count of +2 - +3 is stored in arr[4+2][3+4]
	        z=5                                          # count of +2 - -3 is stored in arr[4+2][-3+4]
	        for i in range(8) :                          # count of +2 - +4 is stored in arr[4+2][4+4]
	            if(i%2==0) :                             # count of +2 - -4 is stored in arr[4+2][-4+4]
	              df.loc[i,"+2 "]=arr[z][6]
	            if(i%2!=0) :
	               z=8-z
	               df.loc[i,"+2 "]=arr[z][6]              # count of +3 - +1 is stored in arr[4+3][1+4]
	               z=8-z                                 # count of +3 - -1 is stored in arr[4+3][-1+4]
	               z=z+1                                 # count of +3 - +2  is stored in arr[4+3][2+4]
	        z=5                                          # count of +3 - -2  is stored in arr[4+3][-2+4]
	        for i in range(8) :                          # count of +3 - +3  is stored in arr[4+3][3+4]
	            if(i%2==0) :                             # count of +3 - -3  is stored in arr[4+3][-3+4]
	              df.loc[i,"-2 "]=arr[z][2]               # count of +3 - +4  is stored in arr[4+3][4+4]
	            if(i%2!=0) :                             # count of +3 - -4  is stored in arr[4+3][-4+4]
	               z=8-z
	               df.loc[i,"-2 "]=arr[z][2]
	               z=8-z
	               z=z+1

	        z=5                                          # count of +4 - +1  is stored in arr[4+4][1+4]
	        for i in range(8) :                          # count of +4 - -1  is stored in arr[4+4][-1+4]
	            if(i%2==0) :                             # count of +4 - +2  is stored in arr[4+4][2+4]
	              df.loc[i,"+3 "]=arr[z][7]               # count of +4 - -2  is stored in arr[4+4][-2+4]
	            if(i%2!=0) :                             # count of +4 - +3  is stored in arr[4+4][3+4]
	               z=8-z                                 # count of +4 - -3  is stored in arr[4+4][-3+4]
	               df.loc[ i,"+3 ",]=arr[z][7]              # count of +4 - +4  is stored in arr[4+4][4+4]
	               z=8-z                                 # count of +4 - -4  is stored in arr[4+4][-4+4]
	               z=z+1

	        z=5
	        for i in range(8) :
	            if(i%2==0) :
	              df.loc[i,"-3 "]=arr[z][1]
	            if(i%2!=0) :
	               z=8-z
	               df.loc[i,"-3 "]=arr[z][1]
	               z=8-z
	               z=z+1
	        z=5
	        for i in range(8) :
	            if(i%2==0) :
	              df.loc[i,"+4 "]=arr[z][8]
	            if(i%2!=0) :
	               z=8-z
	               df.loc[i,"+4 "]=arr[z][8]
	               z=8-z
	               z=z+1
	        z=5
	        for i in range(8) :
	            if(i%2==0) :
	              df.loc[i,"-4 "]=arr[z][0]
	            if(i%2!=0) :
	               z=8-z
	               df.loc[i,"-4 "]=arr[z][0]
	               z=8-z
	               z=z+1

	        bha=0
	        k=14
	        for il in range(d) :
	            rows, cols = (9, 9)
	            arr = [[0 for i in range(cols)] for j in range(rows)]
	            for j in range(mod) :
	                arr[int(df["Octant"][bha])+4][int(df["Octant"][bha+1])+4]=arr[int(df["Octant"][bha])+4][int(df["Octant"][bha+1])+4]+1
	                bha=bha+1
	                if bha>=19996 :
	                    break
	                z=5
	            for ia in range(8) :
	                    if(ia%2==0) :
	                       df.loc[ia+k,"+1 "]=arr[z][5]
	                    if(ia%2!=0) :                   # counting +1 with all other values(-4 to 4) in all modes of transition count
	                       z=8-z
	                       df.loc[ia+k,"+1 "]=arr[z][5]
	                       z=8-z
	                       z=z+1
	            z=5
	            for iy in range(8) :
	                    if(iy%2==0) :
	                       df.loc[iy+k,"-1 "]=arr[z][3]
	                    if(iy%2!=0) :                      # counting -1 with all other values(-4 to 4) in all modes of transition count
	                       z=8-z
	                       df.loc[iy+k,"-1 "]=arr[z][3]
	                       z=8-z
	                       z=z+1
	            z=5
	            for iy1 in range(8) :
	                    if(iy1%2==0) :
	                       df.loc[iy1+k,"+2 "]=arr[z][6]
	                    if(iy1%2!=0) :                      # counting +2 with all other values(-4 to 4) in all modes of transition count
	                       z=8-z
	                       df.loc[iy1+k,"+2 "]=arr[z][6]
	                       z=8-z
	                       z=z+1
	            z=5
	            for iy2 in range(8) :
	                    if(iy2%2==0) :
	                       df.loc[iy2+k,"-2 "]=arr[z][2]
	                    if(iy2%2!=0) :                           # counting -2 with all other values(-4 to 4) in all modes of transition count
	                       z=8-z
	                       df.loc[iy2+k,"-2 "]=arr[z][2]
	                       z=8-z
	                       z=z+1
	            z=5
	            for iy3 in range(8) :
	                    if(iy3%2==0) :
	                       df.loc[iy3+k,"+3 "]=arr[z][7]              # counting +3 with all other values(-4 to 4) in all modes of transition count
	                    if(iy3%2!=0) :
	                       z=8-z
	                       df.loc[iy3+k,"+3 "]=arr[z][7]
	                       z=8-z
	                       z=z+1
	            z=5
	            for iy4 in range(8) :
	                    if(iy4%2==0) :
	                       df.loc[iy4+k,"-3 "]=arr[z][1]
	                    if(iy4%2!=0) :                              # counting -3 with all other values(-4 to 4) in all modes of transition count
	                       z=8-z
	                       df.loc[iy4+k,"-3 "]=arr[z][1]
	                       z=8-z
	                       z=z+1
	            z=5
	            for iy5 in range(8) :
	                    if(iy5%2==0) :
	                       df.loc[iy5+k,"+4 "]=arr[z][8]                  # counting +4 with all other values(-4 to 4) in all modes of transition count
	                    if(iy5%2!=0) :
	                       z=8-z
	                       df.loc[iy5+k,"+4 "]=arr[z][8]
	                       z=8-z
	                       z=z+1
	            z=5
	            for iy6 in range(8) :
	                    if(iy6%2==0) :
	                       df.loc[iy6+k,"-4 "]=arr[z][0]
	                    if(iy6%2!=0) :                                  # counting -4 with all other values(-4 to 4) in all modes of transition count
	                       z=8-z
	                       df.loc[iy6+k,"-4 "]=arr[z][0]
	                       z=8-z
	                       z=z+1

	            k=k+13
	        df.loc[0,"                                                             "]=" "
	        df.loc[0,"Octant ##"]="+1"
	        df.loc[1,"Octant ##"]="-1"
	        df.loc[2,"Octant ##"]="+2"    # creating a column Count for(+1,-1,+2,-2,+3,-3,+4,-4)
	        df.loc[3,"Octant ##"]="-2"
	        df.loc[4,"Octant ##"]="+3"
	        df.loc[5,"Octant ##"]="-3"
	        df.loc[6,"Octant ##"]="+4"
	        df.loc[7,"Octant ##"]="-4"

	        a=[]   #creating a array a for counting sequence (+1,+1 -1,-1 ........ +4,+4 -4,-4)
	        for i in range(9) :
	           a.append(0)    # inserting all elements to 0's in a a[]array
	        b=[]   # creating b array for storing maximum length of each subsequence for (+1,+1 -1,-1 ........ +4,+4 -4,-4)
	        for i in range(9) :
	           b.append(0)     # inserting all elements to 0's in a b[]array
	        c=[]    # creating c array for storing count of for maximum length of each subsequence for (+1,+1 -1,-1 ........ +4,+4 -4,-4)
	        for i in range(9) :
	           c.append(0)  # inserting all elements to 0's in a c[]array
	        mod=19996
	        for i in range(mod) :
	           a[int(df["Octant"][i])+4]=a[int(df["Octant"][i])+4]+1 # incrementing count for each part of subsequene
	           if a[int(df["Octant"][i])+4] != a[int(df["Octant"][i+1])+4]:
	              if(a[int(df["Octant"][i])+4]>b[int(df["Octant"][i])+4]):
	                b[int(df["Octant"][i])+4]=max(a[int(df["Octant"][i])+4],b[int(df["Octant"][i])+4])#finding max length of each subsequence
	                c[int(df["Octant"][i])+4]=1
	                a[int(df["Octant"][i])+4] =0
	              else :
	                if(a[int(df["Octant"][i])+4]==b[int(df["Octant"][i])+4]):
	                   c[int(df["Octant"][i])+4] += 1 # incrementing count for each subsequence of maximum length
	                   a[int(df["Octant"][i])+4] =0
	                if (a[int(df["Octant"][i])+4]<b[int(df["Octant"][i])+4]):
	                    a[int(df["Octant"][i])+4] =0
	        df.loc[0,"Longest Subsequence Length"] = b[5]           # storing maximum length of subsequence +1 in b[5]
	        df.loc[1,"Longest Subsequence Length"] = b[3]           # storing maximum length of subsequence -1 in b[3]
	        df.loc[2,"Longest Subsequence Length"] = b[6]           # storing maximum length of subsequence +2 in b[6]
	        df.loc[3,"Longest Subsequence Length"] = b[2]           # storing maximum length of subsequence -2 in b[2]
	        df.loc[4,"Longest Subsequence Length"] = b[7]           # storing maximum length of subsequence +3 in b[7]
	        df.loc[5,"Longest Subsequence Length"] = b[1]           # storing maximum length of subsequence -3 in b[2]
	        df.loc[6,"Longest Subsequence Length"] = b[8]           # storing maximum length of subsequence +4 in b[8]
	        df.loc[7,"Longest Subsequence Length"] = b[0]           # storing maximum length of subsequence -4 in b[0]

	        df.loc[0,"Count"] = c[5]                                # storing count of maximum length of subsequence +1 in c[5]
	        df.loc[1,"Count"] = c[3]                                # storing count of maximum length of subsequence -1 in c[3]
	        df.loc[2,"Count"] = c[6]                                # storing count of maximum length of subsequence +2 in c[6]
	        df.loc[3,"Count"] = c[2]                                # storing count of maximum length of subsequence -2 in c[2]
	        df.loc[4,"Count"] = c[7]                                # storing count of maximum length of subsequence +3 in c[7]
	        df.loc[5,"Count"] = c[1]                                # storing count of maximum length of subsequence +2 in c[1]
	        df.loc[6,"Count"] = c[8]                                # storing count of maximum length of subsequence +2 in c[8]
	        df.loc[7,"Count"] = c[0]
	        df.loc[0,"   "]=""
	        v=1
	        w=0
	        for i in range(8) :
	            df.loc[w,"Octant ###"]=df["Octant ##"][i]#copying data in octant to OCTANT column
	            df.loc[w,"longest subsequence length"]=df["Longest Subsequence Length"][i] #copying data in Longest subsequence length  to longest subsequence length column
	            df.loc[w,"Frequency"]=df["Count"][i]#copying data in count to Frequency column
	            df.loc[v,"Octant ###"]="Time"#inserting T in OCTANT Column for each value of v
	            df.loc[v,"longest subsequence length"]="From"#inserting FROM in longest subsequence length Column for each value of v
	            df.loc[v,"Frequency"]="To"#inserting To in Frequency  Column for each value of v
	            v=v+2+df["Count"][i]
	            w=w+2+df["Count"][i]
	        rows, cols = (9, 4)
	        begin = [[0 for i in range(cols)] for j in range(rows)]#assiging 0'sto each and every row and column of 2d list begin
	        rows, cols = (9, 4)
	        end = [[0 for i in range(cols)] for j in range(rows)]#assiging 0'sto each and every row and column of 2d list end
	        arr=[]
	        for i in range(9) :
	           arr.append(0)    #assiging 0'sto each element 1d list arr
	        empty=[]
	        for i in range(9) : #assiging 0'sto each element 1d list empty
	           empty.append(0)
	        begin[int(df["Octant"][0])+4][empty[int(df["Octant"][0])+4]]=float(df["T"][0])#assigning begin to 1st value of T coloumn
	        for i in range(19997) :
	            arr[int(df["Octant"][i])+4]=arr[int(df["Octant"][i])+4]+1#incrementing count of each value of octant
	            if(i==19996) :
	              if arr[int(df["Octant"][i])+4]!=b[int(df["Octant"][i])+4] :
	                 begin[int(df["Octant"][i])+4][empty[int(df["Octant"][i])+4]]=0
	                 continue
	            if df["Octant"][i]!=df["Octant"][i+1] :
	                   if arr[int(df["Octant"][i])+4]==(b[int(df["Octant"][i])+4]) :#count of each octant is equal to max of count of each octant
	                     end[int(df["Octant"][i])+4][empty[int(df["Octant"][i])+4]]=float(df["T"][i])#updating end to lqst value of particular octant
	                     empty[int(df["Octant"][i])+4]= empty[int(df["Octant"][i])+4]+1
	                     arr[int(df["Octant"][i])+4]=0
	                     begin[int(df["Octant"][i+1])+4][empty[int(df["Octant"][i+1])+4]]=float(df["T"][i+1])
	                   else :
	                     begin[int(df["Octant"][i])+4][empty[int(df["Octant"][i])+4]]=0#updating begin to 0
	                     arr[int(df["Octant"][i])+4]=0
	                     begin[int(df["Octant"][i+1])+4][empty[int(df["Octant"][i+1])+4]]=float(df["T"][i+1])
	        u=2
	        j1=0
	        for i in range(8) :
	           for j in range(int(df["Count"][j1]))  :
	              df.loc[u,"longest subsequence length"]=begin[int(df["Octant ##"][i])+4][j]
	              df.loc[u,"Frequency"]=end[int(df["Octant ##"][i])+4][j]
	              u=u+1
	           u=u-df["Count"][j1]
	           u=u+2+df["Count"][i]
	           j1=j1+1
	        df.U1 = df.U1.round(3)
	        df.V2 = df.V2.round(3)
	        df.W3 = df.W3.round(3)
	        inp= my_files[sat].replace('.xlsx'," cm_vel_octant_analysis_mod_"+str(check)+".xlsx")
	        # path_out = x+'\output'
	        os.chdir(path_out)
	        print(inp)
	        df.to_excel(inp, index=False)

	        #df.to_excel('2.0.xlsx',index=False)
	        wb=openpyxl.load_workbook(inp)
	        ws=wb['Sheet1']
	        top = Side(border_style='thin',color='000000')
	        bottom = Side(border_style='thin',color='000000')
	        left = Side(border_style='thin',color='000000')
	        right = Side(border_style='thin',color='000000')
	        #border is border for the cell.
	        border = Border(top=top,bottom=bottom,left=left,right=right)
	        range1=ws["N2":"N6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["O2":"O6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["P2":"P6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["Q2":"Q6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["R2":"R6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["S2":"S6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["T2":"T6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["U2":"U6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["V2":"V6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["W2":"W6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["X2":"X6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["Y2":"Y6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["Z2":"Z6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AA2":"AA6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AB2":"AB6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AC2":"AC6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AD2":"AD6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AE2":"AE6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AF2":"AF6"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AI2":"AI9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AJ2":"AJ9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AK2":"AK9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AL2":"AL9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AM2":"AM9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AN2":"AN9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AO2":"AO9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AP2":"AP9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AQ2":"AQ9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AS2":"AS9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AT2":"AT9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AU2":"AU9"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AW2":"AW25"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AX2":"AX25"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["AY2":"AY25"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["P11":"P19"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["Q11":"Q19"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        range1=ws["R11":"R19"]
	        for cell in range1:
	            for x in cell:
	              x.border=border
	        q=15
	        for v in range(d):

	           range1=ws["AI"+str(q):"AI"+str(q+8)]
	           for cell in range1:
	             for x in cell:
	               x.border=border
	           q=q+13
	        q=15
	        for v in range(d):

	           range1=ws["AJ"+str(q):"AJ"+str(q+8)]
	           for cell in range1:
	             for x in cell:
	               x.border=border
	           q=q+13
	        q=15
	        for v in range(d):

	           range1=ws["AK"+str(q):"AK"+str(q+8)]
	           for cell in range1:
	             for x in cell:
	               x.border=border
	           q=q+13
	        q=15
	        for v in range(d):

	           range1=ws["AL"+str(q):"AL"+str(q+8)]
	           for cell in range1:
	             for x in cell:
	               x.border=border
	           q=q+13
	        q=15
	        for v in range(d):

	           range1=ws["AM"+str(q):"AM"+str(q+8)]
	           for cell in range1:
	             for x in cell:
	               x.border=border
	           q=q+13
	        q=15
	        for v in range(d):

	           range1=ws["AN"+str(q):"AN"+str(q+8)]
	           for cell in range1:
	             for x in cell:
	               x.border=border
	           q=q+13
	        q=15
	        for v in range(d):

	           range1=ws["AO"+str(q):"AO"+str(q+8)]
	           for cell in range1:
	             for x in cell:
	               x.border=border
	           q=q+13
	        q=15
	        for v in range(d):

	           range1=ws["AP"+str(q):"AP"+str(q+8)]
	           for cell in range1:
	             for x in cell:
	               x.border=border
	           q=q+13
	        q=15
	        for v in range(d):

	           range1=ws["AQ"+str(q):"AQ"+str(q+8)]
	           for cell in range1:
	             for x in cell:
	               x.border=border
	           q=q+13
	        fill_cell4 = PatternFill(patternType='solid', fgColor='FFFF00')
	        for i in range(d+1):
	            if df["Rank Octant 4"][i]==1 :
	              ws["AC"+str(i+2)].fill = fill_cell4
	        for i in range(d+1):
	            if df["Rank Octant -4"][i]==1 :
	              ws["AD"+str(i+2)].fill = fill_cell4
	        for i in range(d+1):
	            if df["Rank Octant 1"][i]==1 :
	              ws["W"+str(i+2)].fill = fill_cell4
	        for i in range(d+1):
	            if df["Rank Octant -1"][i]==1 :
	              ws["X"+str(i+2)].fill = fill_cell4
	        for i in range(d+1):
	            if df["Rank Octant 2"][i]==1 :
	              ws["Y"+str(i+2)].fill = fill_cell4
	        for i in range(d+1):
	            if df["Rank Octant -2"][i]==1 :
	              ws["Z"+str(i+2)].fill = fill_cell4
	        for i in range(d+1):
	            if df["Rank Octant 3"][i]==1 :
	              ws["AA"+str(i+2)].fill = fill_cell4
	        for i in range(d+1):
	            if df["Rank Octant -3"][i]==1 :
	              ws["AB"+str(i+2)].fill = fill_cell4
	        sat=sat+1
	        wb.save(inp)
	        wb.close()
	    except:
	        print("skipped due error"," cm_vel_octant_analysis_mod_5000.xlsx")
	        continue


mod = 5000
octant_analysis(mod)

# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
