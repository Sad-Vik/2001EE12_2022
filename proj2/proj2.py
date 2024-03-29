
import pandas as pd
import openpyxl
import math
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills
import glob
import os
from datetime import datetime
import streamlit as st
pd.io.formats.excel.ExcelFormatter.header_style = None


start_time = datetime.now()

# @st.cache(suppress_st_warning=True)
def colour_table(inp,l):
    wb = openpyxl.load_workbook(inp)
    ws = wb['Sheet1']
    thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
    )

    fill_cell = PatternFill(fill_type=fills.FILL_SOLID,
                            start_color='00FFFF00', end_color='00FFFF00')

    # define size of the table
    row_num = math.ceil(l/mod)+2
    col_num = 19
    # location of the Table
    row_loc = 1
    col_loc = 14

    for i in range(row_loc, row_loc+row_num):  # for mod
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border
            if ((ws.cell(row=i, column=j).value == 1)):
                ws.cell(row=i, column=j).fill = fill_cell

    # define size of the table
    row_num = 9
    col_num = 3
    # location of the Table
    row_loc = 10
    col_loc = 15

    for i in range(row_loc, row_loc+row_num):  # for mod
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border

    # define size of the table
    row_num = 9
    col_num = 3
    # location of the Table
    row_loc = 1
    col_loc = 45

    for i in range(row_loc, row_loc+row_num):  # for mod
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border

    # define size of the table
    row_num = 9
    col_num = 9
    # location of the Table
    row_loc = 1
    col_loc = 35

    # Number of Tables
    Table_num = math.ceil(l/mod)+1
    dis = 5  # distance between the tables

    for _ in range(Table_num):
        k = 0
        for i in range(row_loc, row_loc+row_num):
            if (i > row_loc):
                ws.cell(row=i, column=col_loc+k).fill = fill_cell
            for j in range(col_loc, col_num+col_loc):
                ws.cell(row=i, column=j).border = thin_border
            k += 1
        row_loc = row_loc+row_num+dis

    row_num = l
    col_num = 3
    # location of the Table
    row_loc = 1
    col_loc = 49

    for i in range(row_loc, row_loc+row_num):  # for mod
        # print(ws.cell(row=i, column=50).value)
        if (ws.cell(row=i, column=50).value == " "):
            break
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border
    wb.save(inp)

# @st.cache(suppress_st_warning=True)
def input_(path):
    os.chdir(path)
    lst_files = glob.glob('*.xlsx')
    return lst_files


# @st.cache(suppress_st_warning=True)
def body_(df1, mod):
    # again we changing the dir to input from output
    # os.chdir()
    # df1 = pd.read_excel(file)  # reading the input file
    avg_u = df1['U'].mean()  # Calculating average of U,V,W
    avg_v = df1['V'].mean()
    avg_w = df1['W'].mean()

    df1["U_Avg"] = ''  # Creating average for coloumns U,V,W
    df1["V_Avg"] = ''
    df1["W_Avg"] = ''
    # assigning the values to respectivley Coloumn
    df1.iloc[0, 4] = round(avg_u, 3)
    df1.iloc[0, 5] = round(avg_v, 3)
    df1.iloc[0, 6] = round(avg_w, 3)

    # Creating new coloumns with Header U',V',W'
    df1["U'=U - U avg"] = round(df1["U"]-avg_u, 3)
    df1["V'=V - V avg"] = round(df1["V"]-avg_v, 3)
    df1["W'=W - W avg"] = round(df1["W"]-avg_w, 3)

    #######          Data PreProcessing     ###########
    df1["Octant"] = ''  # Creatig a empty Column with Header as Octant

    l = len(df1)  # length of DataFrame = 29745
    # creating octant column ,and Identifying the octant value for each triple(U_1,V_1,W_1)
    for i in range(0, l):

        if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] >= 0):
            df1.loc[i, "Octant"] = "+1"  # for +1

        if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] < 0):
            df1.loc[i, "Octant"] = "-1"  # for -1

        if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] >= 0):
            df1.loc[i, "Octant"] = "+2"  # for +2

        if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] < 0):
            df1.loc[i, "Octant"] = "-2"  # for -2

        if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] >= 0):
            df1.loc[i, "Octant"] = "+3"  # for +3

        if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] < 0):
            df1.loc[i, "Octant"] = "-3"  # for -3

        if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] >= 0):
            df1.loc[i, "Octant"] = "+4"  # for +4

        if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] < 0):
            df1.loc[i, "Octant"] = "-4"  # for -4

            ######  Octant Identification  ########

    # creating empty Column without header and assigned "User input" to row 3
    df1[""] = "  "
    df1[" "] = " "
    df1.iloc[0, 12] = "Mod "+str(mod)

    # creating a Coloumn with header as Octant ID
    df1["Octant ID"] = " "
    df1.loc[0, "Octant ID"] = "Overall Octant"

    # oct_count stores a count of unique elements i.e. count of +1,-1,+2,-2,+3,-4,+4
    oct_count = df1['Octant'].value_counts()

    arr = ["+1", "-1", "+2", "-2", "+3", "-3",
           "+4", "-4"]  # cretaed for reference

    oct_cnt = {}  # for storing octant count as key and and coreesponding octant value as value in dict
    for i in range(8):
        s = arr[i]
        # appending the overall count of octant and octant value in dict i.e for Ex:(2610,"+1")
        oct_cnt.update({s: oct_count[s]})
        # And assigning a count values to respectively Coloumns
        df1.loc[0, s] = oct_count[s]

    # sorting the dict by keys
    sortedbyval = {k: v for k, v in sorted(
        oct_cnt.items(), key=lambda item: item[1])}
    # storing the sorted values in a list
    sortedbykey_lst = list(sortedbyval.keys())

    octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                              "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}

    df1["Rank 1"] = ''  # created empty columns
    df1["Rank 2"] = ''
    df1["Rank 3"] = ''
    df1["Rank 4"] = ''
    df1["Rank 5"] = ''
    df1["Rank 6"] = ''
    df1["Rank 7"] = ''
    df1["Rank 8"] = ''
    df1["Rank1 Octant ID"] = " "

    dic_rank = {"+1": "Rank 1", "-1": "Rank 2", "+2": "Rank 3", "-2": "Rank 4",
                "+3": "Rank 5", "-3": "Rank 6", "+4": "Rank 7", "-4": "Rank 8"}  # for reference

    # i=0
    for i in range(8):
        df1.loc[0, dic_rank[sortedbykey_lst[i]]] = 8 - \
            i  # appending the octant ranks of octants
        if (8-i == 1):
            # appending the highest rank octant and its corresponding octant name
            df1.loc[0, "Rank1 Octant ID"] = sortedbykey_lst[i]
            df1.loc[0, "Rank1 Octant Name"] = octant_name_id_mapping[str(
                int(df1.loc[0, "Rank1 Octant ID"]))]

            ###########   Added Some Columns And Rows for MOD Count   ##########

    x = 0  # for findind octant values for MOD ranges
    t = 1  # for row pointer

    count_rank_mod = [0]*8  # Count of rank mod values
    while (x < l):

        d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
              "-3": 5, "+4": 6, "-4": 7}  # creating a dictionary for reference

        # count values of each octant is stored for MOD ranges
        oct_cnt_mod = [0]*8

        oct_cnt = {}  # for storing octant count as key and and coreesponding octant value as value in dict
        for i in range(x, x+mod, 1):

            if (i >= l):
                break  # bound check
            s3 = df1.at[i, "Octant"]
            # incrementing by one of count values of corresponding octants
            oct_cnt_mod[d1[s3]] += 1

        i = 0
        for i in range(8):
            s = arr[i]
            # assigning overall count of octants in each interval
            df1.loc[t, s] = oct_cnt_mod[i]
            # appending the overall count of octant and octant value in dict
            oct_cnt.update({s: oct_cnt_mod[i]})
            # if(i==1):
            # print(oct_cnt_mod[i])
            # print(arr[i])
            # print("hi")

        # sorting the dict by keys
        sortedbyval = {k: v for k, v in sorted(
            oct_cnt.items(), key=lambda item: item[1])}
        # storing the sorted values in a list
        sortedbykey_lst = list(sortedbyval.keys())

        i = 0
        for i in range(8):
            df1.loc[t, dic_rank[sortedbykey_lst[i]]] = 8 - \
                i  # appending the octant ranks of octants
            if (8-i == 1):
                # appending the highest rank octant and its corresponding octant name
                df1.loc[t, "Rank1 Octant ID"] = sortedbykey_lst[i]
                df1.loc[t, "Rank1 Octant Name"] = octant_name_id_mapping[str(
                    int(df1.loc[t, "Rank1 Octant ID"]))]
                # incrementing by one of corresponding octant
                count_rank_mod[d1[sortedbykey_lst[i]]] += 1

        if ((x+mod) > l):  # Writing MOD ranges in Octant ID Coloumn
            df1.loc[t, "Octant ID"] = str(
                x)+"-"+str(l-1)  # for last index(i.e) 2744
        else:
            df1.loc[t, "Octant ID"] = str(x)+"-"+str(x+mod-1)

        x += mod
        t += 1

        ################ Octant Count Based on Mod Values  ######################

    t += 3
    df1.loc[t, "+1"] = "Octant ID"
    df1.loc[t, "-1"] = "Octant Name"
    df1.loc[t, "+2"] = "Count of Rank1 of Mod Values"
    t += 1
    i = 0
    for ID, name in octant_name_id_mapping.items():  # iterating through a dict
        # appending the Octant IDs, Octant Name ,and count of Rank1 of mod values
        df1.loc[t, "+1"] = int(ID)
        df1.loc[t, "-1"] = name
        df1.loc[t, "+2"] = count_rank_mod[i]
        t += 1
        i += 1

        ################ Octant Count Based on Mod Values  ######################
    df1["  "] = ""
    df1["   "] = " "
    df1.iloc[0, 33] = "From"
    df1.loc["Octant #"] = " "
    arr = [" +1", " -1", " +2", " -2", " +3", " -3", " +4", " -4"]

    j = 0
    for i in range(0, 8):
        df1.loc[i, "Octant #"] = arr[j]  # updating Octant ID column
        j += 1

    j = 0
    for j in range(0, 8):
        s1 = arr[j]  # verifing the count of octants
        df1[s1] = " "

    t1 = 0
    t2 = 1
    d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3,
          "+3": 4, "-3": 5, "+4": 6, "-4": 7}
    d2 = {"+1": " +1", "-1": " -1", "+2": " +2", "-2": " -2",
          "+3": " +3", "-3": " -3", "+4": " +4", "-4": " -4"}

    while (1):
        if (t2 == l):
            break
        s1 = df1.at[t1, "Octant"]  # From
        s2 = df1.at[t2, "Octant"]  # To
        # print(df1.loc[d1[s1], d2[s2]])
        if (df1.loc[d1[s1], d2[s2]] == " "):  # checking if cell is empty/null
            df1.loc[d1[s1], d2[s2]] = 1  # adding one
        else:
            # increamenting the count by one and updating it to coloumn
            df1.loc[d1[s1], d2[s2]] = int(df1.loc[d1[s1], d2[s2]]) + 1
        t1 += 1
        t2 += 1

    t = 7
    x = 0
    while (x < l):
        t += 4
        df1.loc[t, "Octant #"] = "Mod Transition Count"
        if ((x+mod) > l):  # Writing MOD ranges in Octant ID Coloumn
            # for last index(i.e) 2744
            df1.loc[t+1, "Octant #"] = str(x)+"-"+str(l-1)
        else:
            df1.loc[t+1, "Octant #"] = str(x)+"-"+str(x+mod-1)
        df1.loc[t+1, " +1"] = "To"
        t += 2
        arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
        df1.loc[t, "Octant #"] = "Octant #"
        df1.iloc[t+1, 33] = "From"
        h = df1.columns  # h stores column labels
        # header name in index format(integer) (here ,y=13)
        y = h.get_loc(" +1")
        j = 0
        for i in range(y, y+8):  # updating a row
            df1.iloc[t, i] = arr[j]
            j += 1

        j = 0
        for i in range(t+1, t+9):  # updating Coloumn
            df1.loc[i, "Octant #"] = arr[j]
            j += 1

        for i in range(x, x+mod):  # each interval

            if (i == l-1):
                break
            s1 = df1.at[i, "Octant"]  # From
            s2 = df1.at[i+1, "Octant"]  # To

            if (df1.loc[t+d1[s1]+1, d2[s2]] == " "):  # checking if cell is empty/null
                df1.loc[t+d1[s1]+1, d2[s2]] = 1  # adding one
            else:
                # increamenting the count by one and updating it to coloumn
                df1.loc[t+d1[s1]+1, d2[s2]
                        ] = int(df1.loc[t+d1[s1]+1, d2[s2]]) + 1
        t += 8
        x += mod

        ##################
    df1["    "] = " "
    df1["Octant ##"] = " "  # Creating empty column with Octant as a header
    arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    for i in range(8):
        # appending values in octant column
        df1.loc[i, "Octant ##"] = arr[i]

    df1["Longest Subsequence Length"] = " "
    df1["Count"] = " "

    x = 0
    # Longest subsequence length for respectively octant values #initlizing a max_count with all zeroes  #initlizing a max_count with all zeroes
    max_count = [0]*8

    # for count of LSL for respectively octant values # initlizing a max_count with all zeroes #initlizing a LSL_count with all zeroes
    LSL_count = [0]*8
    d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
          "-3": 5, "+4": 6, "-4": 7}  # creating a dictionary

    # Creating an empty 2d list of size of 8
    # where each list stores the upper range(Time Range) value of thier respectively Octants
    time_range = []
    for i in range(8):
        time_range.append([])

    while (x < l):
        s1 = df1.at[x, "Octant"]
        count = 0
        j = x
        while (1):  # counting length of sequence
            # breaking if next element is not equal to s1
            if (j >= l or df1.at[j, "Octant"] != s1):
                break
            count += 1
            j += 1

        x += count
        temp = max_count[d1[s1]]

        # updating a maximum count of value if current count is greater the current max
        max_count[d1[s1]] = max(max_count[d1[s1]], count)

        if (count > temp):
            # Reassigning the values of LSL count to one
            LSL_count[d1[s1]] = 1
            # if list is empty appending  Upper range Value
            if (len(time_range[d1[s1]]) == 0):
                time_range[d1[s1]].append(j-1)

            else:
                time_range[d1[s1]].clear()  # Clearing the list
                # appending a curent upper range value to the same clered octant list
                time_range[d1[s1]].append(j-1)

        if (count == temp):
            # incremneting the count of LSL by one
            LSL_count[d1[s1]] += 1
            # appending to the pre-existing(non-empty) list having same LSL of respective Octant
            time_range[d1[s1]].append(j-1)

    for i in range(8):
        # updating Longest subsequence length for respectively octant values
        df1.loc[i, "Longest Subsequence Length"] = max_count[i]

    max_l_cnt = 0
    for j in range(8):
        # updating count of LSL for respectively octant values
        df1.loc[j, "Count"] = LSL_count[j]
        max_l_cnt += LSL_count[j]

    df1["     "] = " "  # Empty Column without Header
    df1["Octant ####"] = " "  # Empty Column
    df1[" Longest Subsequence Length"] = " "  # Empty Column
    df1[" Count"] = " "  # Empty Column
    # print(time_range) # time_range = [[10945], [14645, 18174, 19131], [16990], [29321], [16217], [677], [29219], [28059]]

    t = 0  # row pointer
    for i in range(8):
        df1.loc[t, "Octant ####"] = arr[i]  # Updating Octant Values
        # Updating LSL of Octants
        df1.loc[t, " Longest Subsequence Length"] = max_count[i]
        # updating count of LSl of Octants
        df1.loc[t, " Count"] = LSL_count[i]
        t += 1  # t points to next row
        df1.loc[t, "Octant ####"] = "Time"
        df1.loc[t, " Longest Subsequence Length"] = "From"
        df1.loc[t, " Count"] = "To"

        t += 1  # t points to next row
        for j in range(LSL_count[i]):
            # Appending lower range # From
            df1.loc[t, " Longest Subsequence Length"] = 0.01 * \
                ((time_range[d1[arr[i]]][j])-(max_count[i]-1))
            # Appending Upper range #To
            df1.loc[t, " Count"] = 0.01*time_range[d1[arr[i]]][j]
            t += 1
    return df1 , l



# @st.cache(suppress_st_warning=True)
def output_(lst_files, repeat, data, mod, l, download,i):
    if i == 1:
        inp= lst_files[repeat].replace('.xlsx'," cm_vel_octant_analysis_mod_"+str(mod)+".xlsx")
    if i == 0:
        inp = lst_files.replace(
            '.xlsx', " cm_vel_octant_analysis_mod_"+str(mod)+".xlsx")
    os.chdir(download)
    data.to_excel(inp, index=False)
    wb = openpyxl.load_workbook(inp)
    ws = wb['Sheet1']
    thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000')
                         )

    fill_cell = PatternFill(fill_type=fills.FILL_SOLID,
                            start_color='00FFFF00', end_color='00FFFF00')

    # define size of the table
    row_num = math.ceil(l/mod)+2
    col_num = 19
    # location of the Table
    row_loc = 1
    col_loc = 14

    for i in range(row_loc, row_loc+row_num):  # for mod
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border
            if ((ws.cell(row=i, column=j).value == 1)):
                ws.cell(row=i, column=j).fill = fill_cell

    # define size of the table
    row_num = 9
    col_num = 3
    # location of the Table
    row_loc = math.ceil(l/mod)+6
    col_loc = 15

    for i in range(row_loc, row_loc+row_num):  # for mod
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border

    # define size of the table
    row_num = 9
    col_num = 3
    # location of the Table
    row_loc = 1
    col_loc = 45

    for i in range(row_loc, row_loc+row_num):  # for mod
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border

    # define size of the table
    row_num = 9
    col_num = 9
    # location of the Table
    row_loc = 1
    col_loc = 35

    # Number of Tables
    Table_num = math.ceil(l/mod)+1
    dis = 5  # distance between the tables

    for _ in range(Table_num):
        k = 0
        for i in range(row_loc, row_loc+row_num):
            if (i > row_loc):
                ws.cell(row=i, column=col_loc+k).fill = fill_cell
            for j in range(col_loc, col_num+col_loc):
                ws.cell(row=i, column=j).border = thin_border
            k += 1
        row_loc = row_loc+row_num+dis

    row_num = l
    col_num = 3
    # location of the Table
    row_loc = 1
    col_loc = 49

    for i in range(row_loc, row_loc+row_num):  # for mod
        # print(ws.cell(row=i, column=50).value)
        if (ws.cell(row=i, column=50).value == " "):
            break
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border
    # set_border(ws, 'O')
    wb.save(inp)

st.title("Project 2")
'''
* From *_Mod_* and *_Input Files_*
'''
print(start_time)
genre = st.radio(
    "What is the Input",
    ('None', 'File', 'Folder(Bulk)'))

if genre == 'File':
    # st.write('Lets do this later!.')
    # file = st.text_input("Single File")
    st.subheader("Dataset")
    data_file = st.file_uploader("Upload Excel", type=['xlsx'])
    if data_file is not None:
        file_details = {"Filename": data_file.name,
                            "FileType": data_file.type, "FileSize": data_file.size}
        st.write(file_details)
        lst_files = data_file.name
        mod = st.number_input("Enter Required Mod", min_value=0, step=1)
        out_path = st.text_input("Enter where to download")
        # if not out_path:
        #     out_path = r'C:\Users\dell\Downloads'
        if st.button("Compute !"):
            df = pd.read_excel(data_file)
            data, len = body_(df, mod)
            output_(lst_files,0, data, mod, len, out_path,0)
            os.system('cls')

        # st.dataframe(df)


elif genre == 'Folder(Bulk)':
    path = st.text_input("Enter Folder Path.")
    mod = st.number_input("Enter Required Mod",min_value = 0,step = 1)
    out_path = st.text_input("Enter where to download")
    if not out_path:
        out_path = r'C:\Users\dell\Downloads'
    x = st.button("Compute !")
    if x:
        lst_files = input_(path)
        repeat = 0
        for file in lst_files:
                if file:
                    os.chdir(path)
                    df1 = pd.read_excel(file)
                    df1 , num = body_(df1, mod)
                    output_(lst_files, repeat, df1, mod,num,out_path,1)
                    repeat +=1
        os.system('cls')


end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
st.write("Check the Results.", end_time - start_time)
# mod = 5000
# This shall be the last lines of the code.
